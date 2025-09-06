from typing import Optional, Dict, Any
import os
import hashlib

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from .attachments import get_attachments
try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None  # type: ignore

try:
    # Celery is optional in local dev without worker
    from app.worker.tasks import ingest_from_webhook  # type: ignore
except Exception:  # pragma: no cover
    ingest_from_webhook = None  # type: ignore


class WebhookEvent(BaseModel):
    action: str
    post_id: Optional[int] = None
    url: Optional[str] = None
    meta: Optional[Dict[str, Any]] = None


app = FastAPI(title="etl-api", version="0.1.0")

STORAGE_DIR = os.getenv("STORAGE_DIR", "/data/storage")
UPLOAD_DIR = os.path.join(STORAGE_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)
app.mount("/files", StaticFiles(directory=UPLOAD_DIR), name="files")

# CORS for local dev UIs
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Prometheus metrics
try:
    from prometheus_client import Counter, generate_latest, CONTENT_TYPE_LATEST
    from fastapi.responses import Response

    REQ_COUNTER = Counter("etl_requests_total", "Total requests", ["path"])  # type: ignore

    @app.middleware("http")
    async def _metrics_middleware(request, call_next):  # type: ignore
        response = await call_next(request)
        try:
            REQ_COUNTER.labels(path=request.url.path).inc()  # type: ignore
        except Exception:
            pass
        return response

    @app.get("/metrics")
    def metrics():  # type: ignore
        return Response(generate_latest(), media_type=CONTENT_TYPE_LATEST)
except Exception:
    pass


@app.get("/health")
def health() -> Dict[str, str]:
    return {"status": "ok"}


@app.post("/webhook")
async def webhook(event: WebhookEvent) -> Dict[str, Any]:
    payload = event.model_dump()
    task_id = None
    if ingest_from_webhook:
        async_result = ingest_from_webhook.delay(payload)  # type: ignore
        task_id = async_result.id
    return {"status": "accepted", "task_id": task_id, "event": payload}


@app.post("/ingest/webhook")
async def ingest_webhook(event: WebhookEvent) -> Dict[str, Any]:
    """Alias endpoint per WBS spec."""
    return await webhook(event)


def _sha1_fileobj(fobj) -> str:
    h = hashlib.sha1()
    while True:
        chunk = fobj.read(8192)
        if not chunk:
            break
        h.update(chunk)
    return h.hexdigest()


@app.post("/upload")
async def upload_file(file: UploadFile = File(...)) -> Dict[str, Any]:
    if not file.filename:
        raise HTTPException(status_code=400, detail="filename missing")
    safe_name = os.path.basename(file.filename)
    dest_path = os.path.join(UPLOAD_DIR, safe_name)

    # Write and compute sha1
    with open(dest_path, "wb") as out:
        content = await file.read()
        out.write(content)
    # Reopen for hashing to avoid large memory
    with open(dest_path, "rb") as fin:
        sha1 = _sha1_fileobj(fin)

    public_base = os.getenv("PUBLIC_BASE_URL", "http://localhost:8002")
    internal_base = os.getenv("INTERNAL_BASE_URL", "http://etl-api:8000")
    rel = f"/files/{safe_name}"
    return {
        "filename": safe_name,
        "sha1": sha1,
        "size": os.path.getsize(dest_path),
        "content_type": file.content_type,
        "url": internal_base + rel,
        "public_url": public_base + rel,
    }


@app.get("/posts/{post_id}/attachments")
async def list_post_attachments(post_id: str) -> Dict[str, Any]:
    public_base = os.getenv("PUBLIC_BASE_URL", "http://localhost:8002")
    items = get_attachments(post_id, public_base)
    return {"post_id": post_id, "attachments": items}


@app.get("/preview/xlsx")
async def preview_xlsx(filename: str, sheet: str | None = None, range: str | None = None) -> Dict[str, Any]:
    if not openpyxl:
        raise HTTPException(status_code=500, detail="openpyxl not available")
    path = os.path.join(UPLOAD_DIR, os.path.basename(filename))
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="file not found")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.worksheets[0]
    rng = range
    cells = ws[rng] if rng else ws['A1':'D20']
    data: list[list[str]] = []
    for row in cells:
        data.append(["" if c.value is None else str(c.value) for c in row])
    return {"sheet": ws.title, "range": rng or "A1:D20", "rows": data}
